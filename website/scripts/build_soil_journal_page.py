#!/usr/bin/env python3
"""Build the soil-journal directory from the registry and dated index evidence.

The generated page deliberately separates factual journal descriptions from
indexing evidence. It never converts database coverage into a subjective star
rating or a submission recommendation.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path


CSCD_API = "http://sciencechina.cn/cscdboot/mainPage/getJournalReportAllListPage"
CSCD_PAGE = "http://sciencechina.cn/select"
SCOPUS_PAGE = "https://www.elsevier.com/products/scopus/content"
SCOPUS_FILE = (
    "https://downloads.ctfassets.net/o78em1y1w4i4/7xtaTxNiNcWRTeZkV86eNy/"
    "8df9934a6138c7e15817214c098deaf2/ext_list_Jul_2026.xlsx"
)
CAS_STOP_NOTICE = "https://cssar.cas.cn/library/dtxx/202604/t20260409_8183275.html"
MJL_PAGE = "https://mjl.clarivate.com/collection-list-downloads"
EVIDENCE_DATE = "2026-08-15"


GROUPS = [
    ("中文土壤学及相关专业期刊", 53, "cn-specialist"),
    ("中文综合与交叉期刊", 20, "cn-cross"),
    ("国际土壤学专业期刊", 52, "intl-specialist"),
    ("国际土壤学交叉期刊", 38, "intl-cross"),
    ("国际综合与跨学科期刊", 65, "intl-general"),
]


def esc(value: object) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def normalize(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "").lower()
    value = value.replace("&", "and").replace("（", "(").replace("）", ")")
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", value)


def has_cjk(value: str) -> bool:
    return any("\u4e00" <= char <= "\u9fff" for char in value)


def load_registry(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.DictReader(stream))
    if len(rows) != 228:
        raise ValueError(f"Expected 228 registry rows, found {len(rows)}")
    names = [row["journal_name"].strip() for row in rows]
    if len(names) != len(set(names)):
        raise ValueError("Registry contains duplicate journal names")
    return rows


def fetch_cscd() -> dict[str, dict[str, object]]:
    query = urllib.parse.urlencode(
        {"coreflagC": "C", "coreflagE": "E", "pageNo": 1, "pageSize": 2000}
    )
    with urllib.request.urlopen(f"{CSCD_API}?{query}", timeout=60) as response:
        payload = json.load(response)
    records = payload.get("result", {}).get("records", [])
    if len(records) < 1400:
        raise ValueError(f"CSCD response unexpectedly short: {len(records)} records")
    return {normalize(record["journalName"]): record for record in records}


def load_scopus(path: Path) -> dict[str, list[dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - maintainer-facing message
        raise RuntimeError("openpyxl is required to read the official Scopus XLSX") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Scopus Sources Jul. 2026"]
    iterator = sheet.iter_rows(values_only=True)
    headers = list(next(iterator))
    by_title: dict[str, list[dict[str, object]]] = defaultdict(list)
    for values in iterator:
        record = dict(zip(headers, values))
        title = record.get("Source Title")
        if title:
            by_title[normalize(str(title))].append(record)
    if len(by_title) < 40000:
        raise ValueError(f"Scopus source table unexpectedly short: {len(by_title)} titles")
    return by_title


def registry_group(row: dict[str, str]) -> str:
    name = row["journal_name"]
    group = row["registry_groups"]
    scope = row["soil_topic_scope"]
    if has_cjk(name):
        if group == "Chinese soil and adjacent discovery":
            return "中文土壤学及相关专业期刊"
        return "中文综合与交叉期刊"
    if group == "General and interdisciplinary discovery":
        return "国际综合与跨学科期刊"
    if "直接土壤学" in scope or "Soil Science category" in scope:
        return "国际土壤学专业期刊"
    return "国际土壤学交叉期刊"


def relevance(group_name: str, scope: str) -> tuple[str, str]:
    if group_name == "国际土壤学专业期刊" or "直接土壤学" in scope:
        return "直接土壤", "direct"
    if group_name in {"中文土壤学及相关专业期刊", "国际土壤学交叉期刊"}:
        return "土壤及邻近", "adjacent"
    return "综合交叉", "broad"


def find_cscd(name: str, cscd: dict[str, dict[str, object]]) -> dict[str, object] | None:
    aliases = {
        "中国水土保持科学": "中国水土保持科学（中英文）",
    }
    return cscd.get(normalize(aliases.get(name, name)))


def find_scopus(
    name: str, scopus: dict[str, list[dict[str, object]]]
) -> dict[str, object] | None:
    aliases = {
        "Proceedings of the National Academy of Sciences (PNAS)":
            "Proceedings of the National Academy of Sciences of the United States of America",
        "The ISME Journal": "ISME Journal",
    }
    candidates = scopus.get(normalize(aliases.get(name, name)), [])
    if not candidates:
        return None
    return sorted(
        candidates,
        key=lambda item: (item.get("Active or Inactive") != "Active", str(item.get("Coverage"))),
    )[0]


def enrich(
    rows: list[dict[str, str]],
    cscd: dict[str, dict[str, object]],
    scopus: dict[str, list[dict[str, object]]],
) -> list[dict[str, object]]:
    enriched: list[dict[str, object]] = []
    for row in rows:
        item: dict[str, object] = dict(row)
        group_name = registry_group(row)
        relevance_label, relevance_key = relevance(group_name, row["soil_topic_scope"])
        cscd_record = find_cscd(row["journal_name"], cscd)
        scopus_record = find_scopus(row["journal_name"], scopus)
        cscd_status = str(cscd_record.get("coreFlag")) if cscd_record else ""
        scopus_status = (
            str(scopus_record.get("Active or Inactive")) if scopus_record else ""
        )
        evidence: list[str] = []
        if cscd_status == "C":
            evidence.append("CSCD 2025–2026 核心库（C）")
        elif cscd_status == "E":
            evidence.append("CSCD 2025–2026 扩展库（E）")
        if scopus_status:
            evidence.append(f"Scopus 2026-07 {scopus_status}")

        cautions: list[str] = []
        if has_cjk(row["journal_name"]) and not cscd_status:
            cautions.append("未列入 CSCD 2025–2026 来源期刊表，不能标作当前 CSCD 期刊。")
        if scopus_status == "Inactive":
            coverage = str(scopus_record.get("Coverage") or "未注明")
            cautions.append(f"Scopus 官方来源表标为 Inactive；登记覆盖为 {coverage}。")
        if not row["official_author_url"].strip():
            cautions.append("官方投稿入口尚未完成逐刊核验。")

        if cscd_status == "C":
            assessment = "当前有 CSCD 核心库证据；仍不能替代稿件适配、单位认定或最新投稿规则核验。"
            grade = "核心收录证据"
            grade_key = "strong"
        elif cscd_status == "E":
            assessment = "当前为 CSCD 扩展库，不应与核心库混称；投稿前还应核验单位认定口径。"
            grade = "扩展收录证据"
            grade_key = "moderate"
        elif scopus_status == "Active":
            assessment = "当前 Scopus 来源状态为 Active；这只证明数据库覆盖，不等同于高分区或适合该稿件。"
            grade = "当前收录证据"
            grade_key = "verified"
        elif scopus_status == "Inactive":
            assessment = "当前不应标作 Scopus Active；如仍考虑投稿，需重新核验办刊状态及本单位认可范围。"
            grade = "需谨慎"
            grade_key = "caution"
        else:
            assessment = "尚无本页采用的当前核心收录证据；不得仅凭旧标签或出版社自述作质量结论。"
            grade = "需补充核验"
            grade_key = "caution"

        item.update(
            {
                "group_name": group_name,
                "relevance_label": relevance_label,
                "relevance_key": relevance_key,
                "cscd_status": cscd_status,
                "scopus_status": scopus_status,
                "scopus_coverage": str(scopus_record.get("Coverage") or "") if scopus_record else "",
                "scopus_source_id": str(scopus_record.get("Sourcerecord ID") or "") if scopus_record else "",
                "scopus_oa": str(scopus_record.get("Open Access Status") or "") if scopus_record else "",
                "evidence": evidence,
                "cautions": cautions,
                "assessment": assessment,
                "grade": grade,
                "grade_key": grade_key,
            }
        )
        enriched.append(item)
    return enriched


def write_evidence_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fields = [
        "journal_name",
        "group_name",
        "soil_relevance",
        "publisher",
        "language_or_region",
        "soil_topic_scope",
        "official_author_url",
        "official_source_status",
        "source_accessed_at",
        "cscd_2025_2026",
        "scopus_2026_07",
        "scopus_coverage",
        "scopus_source_id",
        "evidence_grade",
        "caution",
        "evidence_sources",
        "evidence_accessed_at",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "journal_name": row["journal_name"],
                    "group_name": row["group_name"],
                    "soil_relevance": row["relevance_label"],
                    "publisher": row["publisher"],
                    "language_or_region": row["language_or_region"],
                    "soil_topic_scope": row["soil_topic_scope"],
                    "official_author_url": row["official_author_url"],
                    "official_source_status": row["source_status"],
                    "source_accessed_at": row["source_accessed_at"],
                    "cscd_2025_2026": row["cscd_status"],
                    "scopus_2026_07": row["scopus_status"],
                    "scopus_coverage": row["scopus_coverage"],
                    "scopus_source_id": row["scopus_source_id"],
                    "evidence_grade": row["grade"],
                    "caution": " ".join(row["cautions"]),
                    "evidence_sources": f"{CSCD_PAGE} | {SCOPUS_PAGE}",
                    "evidence_accessed_at": EVIDENCE_DATE,
                }
            )


def badge(label: str, kind: str = "neutral") -> str:
    return f'<span class="badge {esc(kind)}">{esc(label)}</span>'


def journal_card(row: dict[str, object]) -> str:
    badges = [badge(str(row["relevance_label"]), f"fit-{row['relevance_key']}")]
    if row["cscd_status"] == "C":
        badges.append(badge("CSCD C · 2025–2026", "good"))
    elif row["cscd_status"] == "E":
        badges.append(badge("CSCD E · 2025–2026", "mid"))
    if row["scopus_status"] == "Active":
        badges.append(badge("Scopus Active · 2026-07", "good"))
    elif row["scopus_status"] == "Inactive":
        badges.append(badge("Scopus Inactive · 2026-07", "warn"))
    if row["official_author_url"]:
        badges.append(badge("官方入口已记录", "source"))
    else:
        badges.append(badge("投稿入口待核验", "warn"))

    evidence_text = "；".join(row["evidence"]) or "本页采用的当前核心收录证据待补充"
    caution_html = ""
    if row["cautions"]:
        caution_html = (
            '<div class="caution"><b>需谨慎</b><span>'
            + esc(" ".join(row["cautions"]))
            + "</span></div>"
        )
    if row["official_author_url"]:
        action = (
            f'<a class="official-link" href="{esc(row["official_author_url"])}" '
            'target="_blank" rel="noreferrer">打开官方投稿入口 ↗</a>'
        )
    else:
        action = '<span class="official-link disabled">官方投稿入口待核验</span>'

    searchable = " ".join(
        str(row[key])
        for key in (
            "journal_name",
            "publisher",
            "language_or_region",
            "soil_topic_scope",
            "group_name",
            "grade",
            "cscd_status",
            "scopus_status",
        )
    ).lower()
    filters = ["all"]
    if row["cscd_status"] == "C":
        filters.append("cscd-c")
    elif row["cscd_status"] == "E":
        filters.append("cscd-e")
    elif has_cjk(str(row["journal_name"])):
        filters.append("cscd-none")
    if row["scopus_status"] == "Active":
        filters.append("scopus-active")
    if row["cautions"]:
        filters.append("review")

    return f"""
            <details class="journal-card" data-search="{esc(searchable)}" data-filters="{esc(' '.join(filters))}">
              <summary>
                <span class="journal-summary-main"><span class="journal-name">{esc(row['journal_name'])}</span><span class="scope-line">适合主题：{esc(row['soil_topic_scope'])}</span></span>
                <span class="summary-grade {esc(row['grade_key'])}">{esc(row['grade'])}</span>
              </summary>
              <div class="journal-body">
                <div class="badges">{''.join(badges)}</div>
                <div class="facts">
                  <div><span>主办 / 出版</span><b>{esc(row['publisher'])}</b></div>
                  <div><span>语种 / 地区</span><b>{esc(row['language_or_region'])}</b></div>
                  <div><span>土壤研究范围</span><b>{esc(row['soil_topic_scope'])}</b></div>
                  <div><span>官方资料状态</span><b>{esc(row['source_status'])} · {esc(row['source_accessed_at'])}</b></div>
                </div>
                <div class="evaluation">
                  <div><span>当前收录证据</span><b>{esc(evidence_text)}</b></div>
                  <div><span>审慎评价</span><p>{esc(row['assessment'])}</p></div>
                </div>
                {caution_html}
                <div class="card-action">{action}<span>证据核验：{EVIDENCE_DATE}</span></div>
              </div>
            </details>"""


def build_html(rows: list[dict[str, object]]) -> str:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["group_name"])].append(row)
    for values in grouped.values():
        values.sort(key=lambda item: str(item["journal_name"]).casefold())

    counts = Counter()
    for row in rows:
        if row["cscd_status"] == "C":
            counts["cscd-c"] += 1
        if row["cscd_status"] == "E":
            counts["cscd-e"] += 1
        if has_cjk(str(row["journal_name"])) and not row["cscd_status"]:
            counts["cscd-none"] += 1
        if row["scopus_status"] == "Active":
            counts["scopus-active"] += 1
        if row["scopus_status"] == "Inactive":
            counts["scopus-inactive"] += 1
        if row["official_author_url"]:
            counts["official"] += 1
        if row["cautions"]:
            counts["review"] += 1

    group_html = []
    for index, (name, expected, key) in enumerate(GROUPS):
        values = grouped[name]
        if len(values) != expected:
            raise ValueError(f"Group {name} expected {expected}, found {len(values)}")
        cards = "".join(journal_card(row) for row in values)
        open_attr = " open" if index == 0 else ""
        group_html.append(
            f"""<details class="journal-group" data-group="{key}"{open_attr}>
          <summary><span>{esc(name)}</span><b><span class="visible-count">{expected}</span> / {expected} 本</b></summary>
          <div class="journal-cards">{cards}
          </div>
        </details>"""
        )

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <meta name="description" content="228本土壤学相关期刊资料库：研究范围、出版社、官方投稿入口、主题适配与带年份的CSCD和Scopus收录证据。">
  <meta name="theme-color" content="#0b5d4b">
  <link rel="canonical" href="https://hemusci.com/skills/soil-journal-format-review/">
  <title>土壤学相关期刊资料库｜228本期刊详情与证据评价</title>
  <style>
    :root{{--ink:#15322c;--muted:#637873;--paper:#f7faf8;--line:rgba(20,78,65,.13);--green:#0f6b56;--deep:#0b3f35;--amber:#9a5a12;--red:#a13c32;--blue:#285e73;--shadow:0 20px 60px rgba(25,67,56,.1)}}
    *{{box-sizing:border-box}}html{{scroll-behavior:smooth}}body{{margin:0;color:var(--ink);background:var(--paper);font:15px/1.72 system-ui,-apple-system,"Segoe UI","Microsoft YaHei","PingFang SC",sans-serif;overflow-x:hidden}}body:before{{content:"";position:fixed;inset:0;pointer-events:none;z-index:-2;background:radial-gradient(circle at 8% 4%,rgba(183,216,107,.23),transparent 28%),radial-gradient(circle at 92% 18%,rgba(46,154,120,.15),transparent 30%),linear-gradient(180deg,#f8fbf7,#f3f8f6 55%,#fbfcf8)}}body:after{{content:"";position:fixed;inset:0;pointer-events:none;z-index:-1;opacity:.2;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='72' height='72' viewBox='0 0 72 72'%3E%3Cg fill='none' stroke='%230f6b56' stroke-opacity='.12'%3E%3Cpath d='M0 36h72M36 0v72'/%3E%3Ccircle cx='36' cy='36' r='1.5' fill='%230f6b56'/%3E%3C/g%3E%3C/svg%3E")}}
    a{{color:inherit}}button,input{{font:inherit}}.wrap{{width:min(1160px,calc(100% - 40px));margin:auto}}nav{{position:sticky;top:0;z-index:20;border-bottom:1px solid rgba(255,255,255,.58);background:rgba(247,250,248,.91);backdrop-filter:blur(18px)}}.nav{{height:72px;display:flex;align-items:center;justify-content:space-between;gap:24px}}.brand{{display:flex;align-items:center;gap:11px;text-decoration:none;font-size:18px;font-weight:850;letter-spacing:-.03em}}.mark{{width:38px;height:38px;display:grid;place-items:center;border-radius:13px;color:white;background:linear-gradient(145deg,var(--green),#143f37)}}.mark svg{{width:23px;height:23px}}.nav-links{{display:flex;align-items:center;gap:22px}}.nav-links a{{color:#45625b;text-decoration:none;font-weight:700}}.nav-links .back{{padding:9px 14px;border-radius:12px;background:var(--green);color:white}}
    .hero{{display:grid;grid-template-columns:1fr auto;gap:48px;align-items:end;padding:68px 0 36px}}.eyebrow{{color:var(--green);font-size:12px;font-weight:850;letter-spacing:.16em;text-transform:uppercase}}.hero h1{{max-width:820px;margin:10px 0 16px;font-size:clamp(40px,5.5vw,64px);line-height:1.08;letter-spacing:-.055em}}.hero p{{max-width:780px;margin:0;color:var(--muted);font-size:18px}}.hero-number{{text-align:right;color:var(--muted)}}.hero-number b{{display:block;color:var(--green);font-size:72px;line-height:.9}}.hero-number span{{font-weight:720}}
    .scope-note{{display:flex;gap:12px;align-items:flex-start;padding:17px 19px;border:1px solid rgba(15,107,86,.13);border-radius:16px;background:rgba(255,255,255,.76);color:#506a64}}.scope-note b{{color:var(--ink);white-space:nowrap}}.stats{{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin:22px 0 34px}}.stat{{padding:17px;border:1px solid var(--line);border-radius:16px;background:rgba(255,255,255,.8)}}.stat b{{display:block;color:var(--green);font-size:25px;line-height:1}}.stat span{{display:block;margin-top:8px;color:var(--muted);font-size:12px}}
    .method{{margin:0 0 28px;padding:26px;border:1px solid var(--line);border-radius:22px;background:rgba(255,255,255,.82);box-shadow:var(--shadow)}}.method-head{{display:grid;grid-template-columns:1fr minmax(280px,.62fr);gap:34px;align-items:start}}.method h2{{margin:5px 0 9px;font-size:26px;letter-spacing:-.035em}}.method p{{margin:0;color:var(--muted)}}.method-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:20px}}.method-item{{padding:15px;border-radius:14px;background:#f1f7f4}}.method-item b{{display:block;margin-bottom:4px}}.method-item span{{color:var(--muted);font-size:13px}}.method-warning{{padding:17px;border-radius:16px;background:#fff6e9;color:#724718}}.method-warning b{{display:block;color:#824d0f}}.method-warning a{{text-underline-offset:3px}}.source-row{{display:flex;flex-wrap:wrap;gap:8px;margin-top:14px}}.source-row a{{padding:7px 10px;border:1px solid var(--line);border-radius:10px;background:white;color:var(--green);font-size:12px;font-weight:760;text-decoration:none}}
    .toolbar{{position:sticky;top:84px;z-index:10;margin:28px 0 16px;padding:14px;border:1px solid var(--line);border-radius:18px;background:rgba(247,250,248,.94);backdrop-filter:blur(14px);box-shadow:0 12px 30px rgba(25,67,56,.08)}}.toolbar-top{{display:grid;grid-template-columns:1fr auto;gap:14px;align-items:center}}.search{{width:100%;padding:13px 15px;border:1px solid var(--line);border-radius:13px;background:white;color:var(--ink)}}.search:focus{{outline:3px solid rgba(15,107,86,.2);border-color:rgba(15,107,86,.35)}}.result{{color:var(--muted);font-weight:760;white-space:nowrap}}.filters{{display:flex;flex-wrap:wrap;gap:8px;margin-top:10px}}.filter{{border:1px solid var(--line);border-radius:999px;padding:7px 11px;background:white;color:#4d6861;cursor:pointer;font-size:12px;font-weight:760}}.filter[aria-pressed="true"]{{border-color:var(--green);background:var(--green);color:white}}
    .journal-groups{{display:grid;gap:12px;padding-bottom:80px}}.journal-group{{overflow:hidden;border:1px solid var(--line);border-radius:18px;background:rgba(255,255,255,.78);box-shadow:0 9px 28px rgba(25,67,56,.04)}}.journal-group>summary{{display:flex;align-items:center;gap:20px;padding:18px 20px;cursor:pointer;list-style:none;font-size:16px;font-weight:810}}.journal-group>summary::-webkit-details-marker,.journal-card>summary::-webkit-details-marker{{display:none}}.journal-group>summary b{{margin-left:auto;color:var(--green);font-size:13px}}.journal-group>summary:after{{content:"＋";color:var(--green)}}.journal-group[open]>summary:after{{content:"－"}}.journal-cards{{display:grid;grid-template-columns:1fr 1fr;gap:10px;padding:0 14px 14px}}
    .journal-card{{border:1px solid var(--line);border-radius:14px;background:#fff;overflow:hidden}}.journal-card>summary{{display:grid;grid-template-columns:1fr auto;gap:14px;align-items:start;min-height:103px;padding:15px;cursor:pointer;list-style:none}}.journal-card>summary:after{{content:"展开";grid-column:2;color:var(--green);font-size:11px;font-weight:760}}.journal-card[open]>summary:after{{content:"收起"}}.journal-summary-main{{min-width:0}}.journal-name{{display:block;font-weight:820;line-height:1.4}}.scope-line{{display:block;margin-top:5px;color:var(--muted);font-size:12px;line-height:1.55}}.summary-grade{{padding:5px 8px;border-radius:9px;font-size:11px;font-weight:820;white-space:nowrap}}.summary-grade.strong,.summary-grade.verified{{background:#e4f4ed;color:#0a664f}}.summary-grade.moderate{{background:#eef3dd;color:#596616}}.summary-grade.caution{{background:#fff0e4;color:#91441b}}.journal-body{{padding:0 15px 16px;border-top:1px solid var(--line)}}.badges{{display:flex;flex-wrap:wrap;gap:6px;padding:13px 0}}.badge{{padding:4px 7px;border-radius:7px;background:#eef3f1;color:#526b65;font-size:10px;font-weight:800}}.badge.good{{background:#e4f4ed;color:#0a664f}}.badge.mid{{background:#edf2d9;color:#586614}}.badge.warn{{background:#fff0e4;color:#97451c}}.badge.source{{background:#eaf2f6;color:#2b5e72}}.badge.fit-direct{{background:#dff1e9;color:#0b644f}}.badge.fit-adjacent{{background:#edf4e7;color:#50691f}}.badge.fit-broad{{background:#edf1f4;color:#49616c}}.facts{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}.facts div,.evaluation>div{{padding:11px;border-radius:11px;background:#f5f8f6}}.facts span,.evaluation span{{display:block;margin-bottom:3px;color:var(--muted);font-size:10px;font-weight:760;text-transform:uppercase;letter-spacing:.04em}}.facts b{{display:block;font-size:12px;line-height:1.55}}.evaluation{{display:grid;grid-template-columns:.85fr 1.15fr;gap:10px;margin-top:10px}}.evaluation b,.evaluation p{{margin:0;font-size:12px;line-height:1.6}}.caution{{display:flex;gap:9px;margin-top:10px;padding:11px;border-radius:11px;background:#fff4e8;color:#764215;font-size:12px}}.caution b{{white-space:nowrap}}.card-action{{display:flex;justify-content:space-between;gap:12px;align-items:center;margin-top:12px}}.official-link{{display:inline-flex;padding:8px 11px;border-radius:10px;background:var(--green);color:white;text-decoration:none;font-size:12px;font-weight:780}}.official-link.disabled{{background:#e8eeeb;color:#6b7d78}}.card-action>span{{color:var(--muted);font-size:10px}}[hidden]{{display:none!important}}
    footer{{padding:38px 0 52px;border-top:1px solid var(--line);color:var(--muted)}}.foot{{display:flex;justify-content:space-between;gap:20px}}.foot a{{color:var(--green);font-weight:740}}
    button:focus-visible,a:focus-visible,summary:focus-visible{{outline:3px solid rgba(15,107,86,.3);outline-offset:2px}}@media(max-width:900px){{.hero{{grid-template-columns:1fr;gap:22px}}.hero-number{{text-align:left}}.hero-number b{{display:inline;margin-right:8px;font-size:42px}}.stats{{grid-template-columns:repeat(3,1fr)}}.method-head{{grid-template-columns:1fr}}.journal-cards{{grid-template-columns:1fr}}}}
    @media(max-width:620px){{.wrap{{width:min(100% - 24px,1160px)}}.nav-links a:not(.back){{display:none}}.hero{{padding-top:48px}}.hero h1{{font-size:38px}}.hero p{{font-size:16px}}.scope-note{{display:block}}.scope-note b{{display:block;margin-bottom:4px}}.stats{{grid-template-columns:1fr 1fr}}.method{{padding:18px}}.method-grid{{grid-template-columns:1fr}}.toolbar{{top:78px}}.toolbar-top{{grid-template-columns:1fr}}.result{{white-space:normal}}.journal-group>summary{{padding:16px}}.journal-cards{{padding:0 9px 9px}}.journal-card>summary{{grid-template-columns:1fr;min-height:0}}.journal-card>summary:after{{grid-column:1}}.summary-grade{{justify-self:start}}.facts,.evaluation{{grid-template-columns:1fr}}.card-action{{align-items:flex-start;flex-direction:column}}.foot{{display:block}}}}
    @media(prefers-reduced-motion:reduce){{html{{scroll-behavior:auto}}}}
  </style>
</head>
<body>
  <nav><div class="wrap nav"><a class="brand" href="/skills/"><span class="mark"><svg viewBox="0 0 24 24" fill="none" aria-hidden="true"><path d="M12 20V9m0 4c-4.6.2-7-2.1-7.2-6.8C9.4 6 11.8 8.3 12 13Zm0 3c4.6.2 7-2.1 7.2-6.8-4.6-.2-7 2.1-7.2 6.8Z" stroke="#efffe7" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"/></svg></span><span>Hemusci Skills</span></a><div class="nav-links"><a href="#method">评价口径</a><a href="#directory">期刊资料库</a><a class="back" href="/skills/">返回技能中心</a></div></div></nav>
  <main class="wrap">
    <header class="hero"><div><div class="eyebrow">Skill 002 · Journal Intelligence</div><h1>土壤学相关期刊<br>详情与证据评价</h1><p>不再只给刊名。每本期刊均展示研究范围、主办或出版社、语种、官方投稿入口、主题适配，以及带版本日期的权威收录证据。</p></div><div class="hero-number"><b>228</b><span>本期刊资料卡</span></div></header>
    <div class="scope-note"><b>边界说明</b><span>本页评价期刊资料与当前数据库收录证据，不对投稿命中率作承诺。排版审查技能仍只处理稿件形式，不评价或改写论文科学内容。</span></div>
    <div class="stats" aria-label="资料库统计"><div class="stat"><b>228</b><span>完整期刊简介</span></div><div class="stat"><b>{counts['official']}</b><span>已记录官方入口</span></div><div class="stat"><b>{counts['cscd-c']}</b><span>CSCD 核心库 C</span></div><div class="stat"><b>{counts['cscd-e']}</b><span>CSCD 扩展库 E</span></div><div class="stat"><b>{counts['scopus-active']}</b><span>Scopus Active</span></div></div>
    <section class="method" id="method">
      <div class="method-head"><div><div class="eyebrow">Evidence, not stars</div><h2>评价看证据，不做主观星级</h2><p>“适不适合土壤研究”和“是否被当前权威数据库收录”分开判断。数据库收录也不等同于期刊分区、论文适配或单位认定。</p></div><div class="method-warning"><b>中科院分区不能写“2026版”</b><span>中国科学院文献情报中心已宣布自 2026 年起不再更新与发布期刊分区表；后续只能注明“2025 最后版”及核验来源。</span></div></div>
      <div class="method-grid"><div class="method-item"><b>官方事实</b><span>主办/出版社、语种、研究范围、投稿入口和核验日期。</span></div><div class="method-item"><b>主题适配</b><span>直接土壤、土壤及邻近、综合交叉；不把跨学科刊误称为土壤专业刊。</span></div><div class="method-item"><b>当前收录</b><span>CSCD 2025–2026 与 Scopus 2026-07；核心库、扩展库、Active/Inactive 分开写。</span></div></div>
      <div class="source-row"><a href="{CSCD_PAGE}" target="_blank" rel="noreferrer">CSCD 官方来源期刊 ↗</a><a href="{SCOPUS_PAGE}" target="_blank" rel="noreferrer">Scopus 官方来源表 ↗</a><a href="{MJL_PAGE}" target="_blank" rel="noreferrer">Web of Science MJL ↗</a><a href="{CAS_STOP_NOTICE}" target="_blank" rel="noreferrer">中科院分区停更声明 ↗</a></div>
    </section>
    <section id="directory">
      <div class="toolbar"><div class="toolbar-top"><label><span hidden>搜索期刊资料</span><input class="search" id="journalSearch" type="search" aria-label="搜索期刊名称、出版社或研究主题" placeholder="搜索刊名、出版社、研究主题或收录状态…" autocomplete="off"></label><span class="result" id="result" aria-live="polite">共 228 本</span></div><div class="filters" role="group" aria-label="期刊证据筛选"><button class="filter" data-filter="all" aria-pressed="true">全部 228</button><button class="filter" data-filter="cscd-c" aria-pressed="false">CSCD 核心 C · {counts['cscd-c']}</button><button class="filter" data-filter="cscd-e" aria-pressed="false">CSCD 扩展 E · {counts['cscd-e']}</button><button class="filter" data-filter="cscd-none" aria-pressed="false">中文未列入 CSCD · {counts['cscd-none']}</button><button class="filter" data-filter="scopus-active" aria-pressed="false">Scopus Active · {counts['scopus-active']}</button><button class="filter" data-filter="review" aria-pressed="false">需复核 · {counts['review']}</button></div></div>
      <div class="journal-groups" id="journalGroups">{''.join(group_html)}</div>
    </section>
  </main>
  <footer><div class="wrap foot"><p>© 2026 Hemusci · 期刊证据核验日期 {EVIDENCE_DATE}</p><p><a href="/skills/#install">安装 soil-journal-format-review →</a></p></div></footer>
  <script>
    const search=document.getElementById('journalSearch');
    const result=document.getElementById('result');
    const groups=[...document.querySelectorAll('.journal-group')];
    const cards=[...document.querySelectorAll('.journal-card')];
    const filters=[...document.querySelectorAll('.filter')];
    let activeFilter='all';
    function applyFilters(){{
      const query=search.value.trim().toLocaleLowerCase();
      let visible=0;
      groups.forEach(group=>{{
        let groupVisible=0;
        group.querySelectorAll('.journal-card').forEach(card=>{{
          const searchMatch=!query||card.dataset.search.includes(query);
          const filterMatch=activeFilter==='all'||card.dataset.filters.split(' ').includes(activeFilter);
          const match=searchMatch&&filterMatch;
          card.hidden=!match;
          if(match){{visible+=1;groupVisible+=1}}
        }});
        group.hidden=groupVisible===0;
        group.querySelector('.visible-count').textContent=groupVisible;
        if((query||activeFilter!=='all')&&groupVisible)group.open=true;
      }});
      result.textContent=(query||activeFilter!=='all')?`找到 ${{visible}} 本`:'共 228 本';
    }}
    search.addEventListener('input',applyFilters);
    filters.forEach(button=>button.addEventListener('click',()=>{{
      activeFilter=button.dataset.filter;
      filters.forEach(item=>item.setAttribute('aria-pressed',String(item===button)));
      applyFilters();
    }}));
  </script>
</body>
</html>
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--registry", type=Path, required=True)
    parser.add_argument("--scopus-xlsx", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--evidence-csv", type=Path, required=True)
    args = parser.parse_args()

    rows = load_registry(args.registry)
    cscd = fetch_cscd()
    scopus = load_scopus(args.scopus_xlsx)
    enriched = enrich(rows, cscd, scopus)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(build_html(enriched), encoding="utf-8")
    write_evidence_csv(args.evidence_csv, enriched)
    print(
        json.dumps(
            {
                "journals": len(enriched),
                "official_urls": sum(bool(row["official_author_url"]) for row in enriched),
                "cscd_c": sum(row["cscd_status"] == "C" for row in enriched),
                "cscd_e": sum(row["cscd_status"] == "E" for row in enriched),
                "scopus_active": sum(row["scopus_status"] == "Active" for row in enriched),
                "scopus_inactive": sum(row["scopus_status"] == "Inactive" for row in enriched),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
